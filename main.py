__author__ = "{A. Venton}"
__copyright__ = "Copyright 2023, AdvancedMD Schedule Generator"
__version__ = "1.0"
__email__ = "{aaventon@gmail.com}"

import csv
import re
from os import path

import PySimpleGUI as sg
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


def write_csv_to_xlsx(csv_filename, ws):
    """
    Write csv file to xlsx file with specified columns and rows

    :param csv_filename: Str of name of csv file
    :param ws: Workbook xlsx object
    :return: ws, provider_name, date_range, practice_name
    """

    # Write csv data to excel files
    with open(csv_filename, encoding="utf-8-sig") as csv_file:

        # Add csv data to a dictionary
        csv_reader = csv.DictReader(csv_file)

        # Variables to get specific header data
        provider_name = ""
        date_range = ""
        practice_name = ""
        iteration = 0
        for csv_row in csv_reader:

            # Capture header data
            if iteration == 0:
                provider_name = get_provider(csv_row["Textbox9"])
                date_range = csv_row["textbox29"]
                practice_name = csv_row["PracticeName"]
                iteration += 1

            # Insert only these values into rows
            patient_appointment = [csv_row["AppointmentTime"],
                                   csv_row["Patient"],
                                   csv_row["Comments"],
                                   csv_row["PatientEmailAddress"],
                                   csv_row["AppointmentTypeName"],
                                   csv_row["Carrier"],
                                   csv_row["Provider"]]

            ws.append(patient_appointment)

    return ws, provider_name, date_range, practice_name


def get_provider(provider_string):
    """
    Helper function to get substring from given string

    :param provider_string: Str that contains provider name
    :return: provider_name
    """

    # String starts with "Providers"
    provider_match = re.search(r"^P.+(,\s)", provider_string)

    if provider_match:
        provider_name = provider_match.group(0)
    else:
        provider_name = provider_string

    return provider_name


def apply_styles(ws, provider_name, date_range, practice_name, font_name, provider_font_size, text_size):
    """
    Apply excel styles to each row in excel sheet

    :param ws: Workbook xlsx object
    :param provider_name: Str of provider name
    :param date_range: Str of date range
    :param practice_name: Str of practice name
    :param font_name: Selected font
    :param provider_font_size: Selected font size for provider
    :param text_size: Selected font size for rest of text
    :return:
    """

    # Create font styles
    provider_font, header_font, row_font = font_styles(font_name, provider_font_size, text_size)

    # Insert header rows
    ws.insert_rows(1, amount=5)

    # Write "Provider" cell
    ws.cell(row=1, column=1, value=provider_name).font = provider_font

    # Write "Date" and "Practice" cells
    ws.cell(row=2, column=1, value=date_range).font = header_font
    ws.cell(row=3, column=1, value=practice_name).font = header_font
    for cell in ws["5:5"]:
        cell.font = header_font

    # Write header text to each column in row 5, A-G
    ws["A5"] = "Time"
    ws["B5"] = "Patient"
    ws["C5"] = "Comments"
    ws["D5"] = "Email"
    ws["E5"] = "Type"
    ws["F5"] = "Carrier"
    ws["G5"] = "Provider"

    # Apply row font and wrap text in "Comments" column
    for row in ws.iter_rows(min_row=6, max_col=8):
        for cell in row:
            cell.font = row_font
            if cell.column == 3:
                cell.alignment = Alignment(wrapText=True)

    return ws


def font_styles(font_name, provider_font_size, text_size):
    """
    Helper function to create excel sheet font style objects

    :param font_name: Selected font
    :param provider_font_size: Selected font size for provider
    :param text_size: Selected font size for rest of text
    :return: provider_font, header_font, row_font
    """

    # Create font styles
    provider_font = Font(name=font_name, bold=True, size=provider_font_size)
    header_font = Font(name=font_name, bold=True, size=text_size)
    row_font = Font(name=font_name, size=text_size)

    return provider_font, header_font, row_font


def apply_print_settings(ws):
    """
    Apply print settings to excel sheet

    :param ws: Workbook xlsx object
    :return: ws
    """
    # Set page layout
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

    # Set print titles
    ws.print_title_rows = "1:5"

    # Set column dimensions for Columns B-D to fit on printed page
    ws.column_dimensions["B"].width = 19
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 19

    return ws


def generate_output_filename(provider_name):
    """
    Generates output file name based on provider
    :param provider_name: Str of provider name
    :return: output_filename
    """

    # Replace special characters with white space
    output_filename = re.sub(r"[\W]+", "_", provider_name)

    return output_filename


def check_path_valid(file_selected):
    """
    Check user input and output file validity
    :param file_selected: Str file path
    :return: none
    """

    # Check that dictionary is not empty and file path exists
    if file_selected and path.exists(file_selected):
        return True
    else:
        sg.popup_error("Enter valid input and output filepath")


if __name__ == "__main__":

    layout = [
        [sg.Text("Input CSV file: "), sg.Input(key="-IN-"), sg.FileBrowse(file_types=(("CSV Files", "*.csv*"),))],
        [sg.Text("Output folder: "), sg.Input(key="-OUT-"), sg.FolderBrowse()],
        [sg.Exit("Exit"), sg.Button("Submit")]
    ]

    window = sg.Window("File Browser", layout)

    while True:
        event, values = window.read()
        if event == sg.WIN_CLOSED or event == "Exit":
            break
        elif event == "Submit":
            # Assign file path and folder path from dict

            file_selected = path.abspath(values["-IN-"])
            folder_selected = path.abspath(values["-OUT-"])

            # Check if file is valid
            if check_path_valid(file_selected) and check_path_valid(folder_selected):

                # Create excel workbook
                wb = Workbook()
                ws = wb.active

                # Write csv data to excel file
                csv_filename = file_selected
                ws, provider_name, date_range, practice_name = write_csv_to_xlsx(csv_filename, ws)

                # Formatting
                font_name = "Tahoma"
                provider_font_size = 14
                text_size = 8
                ws = apply_styles(ws, provider_name, date_range, practice_name, font_name, provider_font_size, text_size)

                # Print settings
                ws = apply_print_settings(ws)

                # Get output filename based on provider
                output_filename = generate_output_filename(provider_name)

                # Save workbook
                wb.save(filename="{}/{}.xlsx".format(folder_selected,output_filename))

                sg.popup_no_titlebar("File generated!")